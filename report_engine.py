import os
from datetime import datetime
from typing import Any, Dict, Optional

import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference


REPORTS_DIR = "reports"


def _ensure_dir(path: str):
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _safe_name(text: str, max_len: int = 60) -> str:
    text = (text or "report").strip().lower()
    for ch in ["/", "\\", ":", "*", "?", '"', "<", ">", "|"]:
        text = text.replace(ch, "")
    text = text.replace(" ", "_")
    return text[:max_len] if len(text) > max_len else text


def _to_dataframe(result: Any) -> Optional[pd.DataFrame]:
    """
    Try to coerce various result types into a DataFrame.
    """
    if result is None:
        return None

    if isinstance(result, pd.DataFrame):
        return result

    if isinstance(result, pd.Series):
        return result.to_frame(name=result.name or "value")

    if isinstance(result, dict):
        try:
            return pd.DataFrame([result])
        except Exception:
            return None

    if isinstance(result, list):
        try:
            return pd.DataFrame(result)
        except Exception:
            return None

    # fallback: scalar → single cell DF
    try:
        return pd.DataFrame({"result": [result]})
    except Exception:
        return None


def _extract_key_metrics(df: Optional[pd.DataFrame]):
    """
    Extract simple key metrics: max/min/mean for numeric columns.
    """
    metrics = []
    if df is None or df.empty:
        return metrics

    num_df = df.select_dtypes(include=["number"])
    if num_df.empty:
        return metrics

    for col in num_df.columns[:3]:  # limit to first 3 numeric columns
        series = num_df[col].dropna()
        if series.empty:
            continue
        metrics.append({
            "metric": col,
            "max": float(series.max()),
            "min": float(series.min()),
            "avg": float(series.mean()),
        })
    return metrics


def _generate_executive_summary(query: str, explanation: str, metrics):
    """
    Create a concise executive summary.
    """
    lines = []
    lines.append(f"Query: {query}")

    if metrics:
        top = metrics[0]
        lines.append(
            f"Key metric '{top['metric']}' ranges from {top['min']:.2f} to {top['max']:.2f} (avg {top['avg']:.2f})."
        )

    if explanation:
        # keep it short
        short = explanation.split("\n")[0][:200]
        lines.append(f"Insight: {short}")

    # simple recommendation heuristic
    if metrics:
        m = metrics[0]
        if m["max"] > m["avg"] * 1.2:
            lines.append("Recommendation: Investigate top performers and replicate drivers.")
        else:
            lines.append("Recommendation: Focus on steady growth and consistency.")

    return " ".join(lines)


def _autosize_worksheet(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


# --- SMART CHART HELPERS ---
def _get_numeric_columns(ws):
    numeric_cols = []
    for col_idx in range(1, ws.max_column + 1):
        is_numeric = False
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if isinstance(val, (int, float)):
                is_numeric = True
                break
        if is_numeric:
            numeric_cols.append(col_idx)
    return numeric_cols


def _get_time_column(ws):
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(row=1, column=col_idx).value).lower()
        if any(k in header for k in ["date", "time", "month", "year"]):
            return col_idx
    return None



# SMART CHART: auto-detect numeric/time columns and plot
def _add_smart_chart(ws, max_rows=20):
    """
    Create smart chart:
    - Line chart if time column exists
    - Otherwise bar chart with first numeric column
    """
    try:
        rows = ws.max_row
        if rows < 2:
            return

        numeric_cols = _get_numeric_columns(ws)
        if not numeric_cols:
            return

        time_col = _get_time_column(ws)

        if time_col:
            # --- Line Chart ---
            chart = LineChart()
            chart.title = "Trend Analysis"

            for col in numeric_cols[:2]:  # max 2 series
                data = Reference(ws, min_col=col, min_row=1, max_row=min(rows, max_rows))
                chart.add_data(data, titles_from_data=True)

            cats = Reference(ws, min_col=time_col, min_row=2, max_row=min(rows, max_rows))
            chart.set_categories(cats)

        else:
            # --- Bar Chart ---
            chart = BarChart()
            chart.title = "Key Metrics"

            col = numeric_cols[0]
            data = Reference(ws, min_col=col, min_row=1, max_row=min(rows, max_rows))
            chart.add_data(data, titles_from_data=True)

            cats = Reference(ws, min_col=1, min_row=2, max_row=min(rows, max_rows))
            chart.set_categories(cats)

        ws.add_chart(chart, "E2")

    except Exception:
        pass


def save_txt_report(
    query: str,
    result: Any,
    explanation: str,
    context: Optional[str] = None,
    out_dir: str = REPORTS_DIR,
) -> str:
    _ensure_dir(out_dir)

    fname = f"{_safe_name(query)}_{_timestamp()}.txt"
    path = os.path.join(out_dir, fname)

    with open(path, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        f.write("AI ANALYST REPORT\n")
        f.write("=" * 70 + "\n\n")

        # --- Executive Summary ---
        df = _to_dataframe(result)
        metrics = _extract_key_metrics(df)
        summary = _generate_executive_summary(query, explanation, metrics)

        f.write("EXECUTIVE SUMMARY\n")
        f.write("-" * 70 + "\n")
        f.write(summary + "\n\n")

        f.write("QUERY\n")
        f.write("-" * 70 + "\n")
        f.write(f"{query}\n\n")

        f.write("KEY RESULTS\n")
        f.write("-" * 70 + "\n")
        f.write(str(result) + "\n\n")

        if metrics:
            f.write("KEY METRICS\n")
            f.write("-" * 70 + "\n")
            for m in metrics:
                f.write(f"{m['metric']}: max={m['max']:.2f}, min={m['min']:.2f}, avg={m['avg']:.2f}\n")
            f.write("\n")

        if explanation:
            f.write("INSIGHTS & EXPLANATION\n")
            f.write("-" * 70 + "\n")
            f.write(explanation + "\n\n")

        if context:
            f.write("CONTEXT\n")
            f.write("-" * 70 + "\n")
            f.write(context + "\n\n")

        f.write(f"Generated at: {datetime.now()}\n")

    return path


def save_excel_report(
    query: str,
    result: Any,
    explanation: str,
    context: Optional[str] = None,
    out_dir: str = REPORTS_DIR,
) -> str:
    _ensure_dir(out_dir)

    fname = f"{_safe_name(query)}_{_timestamp()}.xlsx"
    path = os.path.join(out_dir, fname)

    df = _to_dataframe(result)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Sheet 1: Data
        if df is not None:
            df.to_excel(writer, sheet_name="data", index=False)
        else:
            pd.DataFrame({"result": [str(result)]}).to_excel(
                writer, sheet_name="data", index=False
            )

        # Sheet 2: Summary
        summary_rows = [
            {"Section": "Query", "Value": query},
            {"Section": "Generated At", "Value": str(datetime.now())},
        ]
        if explanation:
            summary_rows.append({"Section": "Explanation", "Value": explanation})
        if context:
            summary_rows.append({"Section": "Context", "Value": context})

        summary_df = pd.DataFrame(summary_rows)
        summary_df.to_excel(writer, sheet_name="summary", index=False)

        # Sheet 3: Executive Summary
        df_local = _to_dataframe(result)
        metrics = _extract_key_metrics(df_local)
        exec_summary = _generate_executive_summary(query, explanation, metrics)

        exec_df = pd.DataFrame({
            "Executive Summary": [exec_summary]
        })
        exec_df.to_excel(writer, sheet_name="executive_summary", index=False)

        # Formatting
        workbook = writer.book

        # Data sheet formatting
        ws_data = workbook["data"]
        for cell in ws_data[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Highlight top rows
        for row in ws_data.iter_rows(min_row=2, max_row=6):
            for cell in row:
                cell.font = Font(bold=True)

        _autosize_worksheet(ws_data)

        # Add chart
        _add_smart_chart(ws_data)

        # Summary sheet formatting
        ws_summary = workbook["summary"]
        for cell in ws_summary[1]:
            cell.font = Font(bold=True)

        for row in ws_summary.iter_rows(min_row=2):
            row[0].font = Font(bold=True)

        _autosize_worksheet(ws_summary)

        # Executive summary formatting
        ws_exec = workbook["executive_summary"]
        for cell in ws_exec[1]:
            cell.font = Font(bold=True)
        _autosize_worksheet(ws_exec)

    return path


def generate_reports(
    query: str,
    result: Any,
    explanation: str,
    context: Optional[str] = None,
    out_dir: str = REPORTS_DIR,
) -> Dict[str, str]:
    """
    Generate both TXT and Excel reports.
    Returns paths.
    """
    txt_path = save_txt_report(query, result, explanation, context, out_dir)
    xlsx_path = save_excel_report(query, result, explanation, context, out_dir)

    return {"txt": txt_path, "xlsx": xlsx_path}